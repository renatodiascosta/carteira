#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lê as colunas H e I da aba "Aportes" da planilha Registro 2026.xlsx
e gera acoes_e_setores.json.

Mapeamento das colunas (linha 1 = cabeçalho, dados a partir da linha 2):
    H -> Ação  -> ticker
    I -> Setor -> setor

Dependência: openpyxl  (pip install openpyxl)
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------- configuração
ORIGEM = Path(r"C:\Users\User\OneDrive\Operações\Registro 2026.xlsx")
DESTINO = Path(r"C:\Users\User\Desktop\Python\carteira\acoes_e_setores.json")
ABA = "Aportes"

COL_TICKER = 8           # H
COL_SETOR = 9            # I
LINHA_INICIAL = 2        # pula o cabeçalho
ORDENAR = False          # True -> ordena alfabeticamente por ticker
CHAVE_RAIZ = None        # None -> JSON é uma lista; ex.: "ativos" -> {"ativos": [...]}


def ler_setores(caminho: Path, aba: str) -> list[dict]:
    # data_only=True devolve o resultado calculado das fórmulas, não o texto da fórmula
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            raise SystemExit(f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        ws = wb[aba]

        registros = []
        vistos = set()
        for n, (col_h, col_i) in enumerate(
            ws.iter_rows(
                min_row=LINHA_INICIAL,
                min_col=COL_TICKER,
                max_col=COL_SETOR,
                values_only=True,
            ),
            start=LINHA_INICIAL,
        ):
            ticker = str(col_h).strip().upper() if col_h is not None else ""
            if not ticker:
                continue  # ignora linhas em branco e o bloco de totais mais abaixo

            setor = str(col_i).strip() if col_i is not None else ""
            if not setor:
                print(f"  aviso: linha {n} ({ticker}) está sem setor", file=sys.stderr)

            if ticker in vistos:
                print(f"  aviso: linha {n} ({ticker}) é duplicada", file=sys.stderr)
            vistos.add(ticker)

            registros.append({"ticker": ticker, "setor": setor})

        if ORDENAR:
            registros.sort(key=lambda r: r["ticker"])
        return registros
    finally:
        wb.close()


def main():
    if not ORIGEM.exists():
        raise SystemExit(f"Planilha não encontrada: {ORIGEM}")

    ativos = ler_setores(ORIGEM, ABA)
    saida = ativos if CHAVE_RAIZ is None else {CHAVE_RAIZ: ativos}

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    with DESTINO.open("w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)

    print(f"{len(ativos)} ativos gravados em {DESTINO}")


if __name__ == "__main__":
    main()