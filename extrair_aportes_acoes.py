#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lê a aba "Aportes" da planilha Registro 2026.xlsx e gera aportes.json.

Mapeamento das colunas (linha 1 = cabeçalho, dados a partir da linha 2):
    A -> Ação        -> ticker
    B -> Qtd.        -> quantidade
    C -> Data        -> data
    D -> Preço médio -> preco_medio

Dependência: openpyxl  (pip install openpyxl)
"""

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------- configuração
ORIGEM = Path(r"C:\Users\User\OneDrive\Operações\Registro 2026.xlsx")
DESTINO = Path(r"C:\Users\User\Desktop\Python\carteira\aportes.json")
ABA = "Aportes"

LINHA_INICIAL = 2        # pula o cabeçalho
FORMATO_DATA = "%Y-%m-%d"  # troque para "%d/%m/%Y" se preferir o formato BR
CASAS_PRECO = 2          # arredondamento do preço médio; use None para não arredondar
CHAVE_RAIZ = None        # None -> JSON é uma lista; ex.: "aportes" -> {"aportes": [...]}


def formatar_data(valor):
    """Converte o valor da célula de data para string."""
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime(FORMATO_DATA)
    return str(valor).strip()


def para_numero(valor, casas=None):
    """Converte a célula para int/float, tolerando texto com vírgula decimal."""
    if valor is None:
        return None
    if isinstance(valor, str):
        valor = valor.strip().replace(".", "").replace(",", ".")
        if not valor:
            return None
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return None
    if casas is not None:
        numero = round(numero, casas)
    if numero.is_integer() and casas in (0, None):
        return int(numero)
    return numero


def ler_aportes(caminho: Path, aba: str) -> list[dict]:
    # data_only=True devolve o resultado calculado das fórmulas, não o texto da fórmula
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            raise SystemExit(f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        ws = wb[aba]

        registros = []
        for n, (col_a, col_b, col_c, col_d) in enumerate(
            ws.iter_rows(min_row=LINHA_INICIAL, min_col=1, max_col=4, values_only=True),
            start=LINHA_INICIAL,
        ):
            ticker = str(col_a).strip().upper() if col_a is not None else ""
            if not ticker:
                continue  # ignora linhas em branco no meio ou no fim da aba

            quantidade = para_numero(col_b, casas=0)
            preco_medio = para_numero(col_d, casas=CASAS_PRECO)
            data_str = formatar_data(col_c)

            if quantidade is None or preco_medio is None or data_str is None:
                print(f"  aviso: linha {n} ({ticker}) tem campo vazio/inválido", file=sys.stderr)

            registros.append(
                {
                    "ticker": ticker,
                    "data": data_str,
                    "quantidade": quantidade,
                    "preco_medio": preco_medio,
                }
            )
        return registros
    finally:
        wb.close()


def main():
    if not ORIGEM.exists():
        raise SystemExit(f"Planilha não encontrada: {ORIGEM}")

    aportes = ler_aportes(ORIGEM, ABA)
    saida = aportes if CHAVE_RAIZ is None else {CHAVE_RAIZ: aportes}

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    with DESTINO.open("w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)

    print(f"{len(aportes)} aportes gravados em {DESTINO}")


if __name__ == "__main__":
    main()